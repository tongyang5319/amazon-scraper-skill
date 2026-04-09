"""
    Main collector: orchestrates list scraping + detail scraping + CSV/XLSX output.
"""

import io
import logging
import os
import re
import time
from datetime import date
from typing import Optional

import pandas as pd
import requests

from amazon_unified_scraper.list_scraper import ListScraper
from amazon_unified_scraper.detail_scraper import DetailScraper
from amazon_unified_scraper.models import EnrichedProduct, ListProduct


class UnifiedCollector:
    """
    High-level orchestrator:
      1. Scrape list page → ListProduct[]
      2. For each product, scrape detail page → DetailProduct
      3. Merge & save to CSV (always) and optionally XLSX with embedded images
    """

    # Fields that trigger auto-retry when missing
    _RETRY_FIELDS = (
        "bought_in_past_month",
        "rating",
        "review_count",
        "sub_category_rank",
        "brand",
        "product_size",
        "product_weight",
    )

    def __init__(
        self,
        output_dir: str = "data",
        delay_range: tuple[float, float] = (8.0, 15.0),
        postal_code: str | None = None,
        output_format: str = "csv",
        auto_retry: bool = True,
        logger: Optional[logging.Logger] = None,
    ) -> None:
        self._list_scraper = ListScraper(postal_code=postal_code)
        self._detail_scraper = DetailScraper(postal_code=postal_code)
        self._output_dir = output_dir
        self._delay_range = delay_range
        self._postal_code = postal_code
        self._output_format = output_format
        self._auto_retry = auto_retry
        self._logger = logger if logger else logging.getLogger(__name__)
        os.makedirs(self._output_dir, exist_ok=True)
        # Image cache dir for XLSX export
        self._img_dir = os.path.join(self._output_dir, ".product_images")
        if output_format == "xlsx":
            os.makedirs(self._img_dir, exist_ok=True)

    def _clean_category_name(self, name: str) -> str:
        """Clean a category name into a safe filename string."""
        if not name or name == "unknown":
            return "unknown_category"
        # Remove special chars, keep Chinese/alphanumeric
        name = re.sub(r"[^\w\s\u4e00-\u9fff\-]", "", name)
        name = re.sub(r"\s+", "_", name.strip())
        return name[:40] or "category"

    def _build_output_path(self, category_name: str, ext: str = "csv") -> str:
        today = date.today().strftime("%Y%m%d")
        safe_cat = self._clean_category_name(category_name)
        filename = f"新品榜_{safe_cat}_{today}.{ext}"
        return os.path.join(self._output_dir, filename)

    def _download_image(self, asin: str, image_url: str) -> str | None:
        """Download product image to cache dir, return local path or None."""
        if not image_url:
            return None
        cache_path = os.path.join(self._img_dir, f"{asin}.jpg")
        if os.path.exists(cache_path):
            return cache_path
        try:
            resp = requests.get(image_url, timeout=10, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            if resp.status_code == 200:
                with open(cache_path, "wb") as f:
                    f.write(resp.content)
                return cache_path
        except Exception:
            pass
        return None

    def _enrich_product(
        self, list_prod: ListProduct, max_reviews: int
    ) -> EnrichedProduct:
        """
        Visit a product's detail page and enrich with detail data.
        Applies anti-detection delays automatically inside DetailScraper.scrape().
        """
        try:
            detail = self._detail_scraper.scrape(list_prod.url, list_prod.asin_code)

            # Build bullet point fields
            bp_fields = {}
            if detail.bullet_points:
                for i, bp in enumerate(detail.bullet_points[:5], 1):
                    bp_fields[f"bullet_point_{i}"] = bp

            # Build review fields (rating + text per review)
            review_fields = {}
            if detail.reviews:
                for i, r in enumerate(detail.reviews[:max_reviews], 1):
                    review_fields[f"review_{i}_rating"] = r.get("rating") or ""
                    review_fields[f"review_{i}_text"] = r.get("text", "")[:300]

            return EnrichedProduct(
                # From list page
                list_rank=list_prod.list_rank,
                title=list_prod.title,
                url=list_prod.url,
                asin_code=list_prod.asin_code,
                image_url=list_prod.image_url,
                price=list_prod.price,
                rating=list_prod.rating or detail.rating,
                review_count=list_prod.review_count or detail.review_count,
                # From detail page
                brand=detail.brand,
                bought_in_past_month=detail.bought_in_past_month,
                has_coupon=detail.has_coupon,
                coupon_text=detail.coupon_text,
                product_size=detail.product_size,
                product_weight=detail.product_weight,
                sub_category_name=detail.sub_category_name,
                sub_category_rank=detail.sub_category_rank,
                has_a_plus=detail.has_a_plus,
                **bp_fields,
                **review_fields,
            )
        except Exception as e:
            self._logger.warning(
                f"  Failed to enrich ASIN {list_prod.asin_code}: {e}"
            )
            # Return list-only data if detail scraping fails
            return EnrichedProduct(
                list_rank=list_prod.list_rank,
                title=list_prod.title,
                url=list_prod.url,
                asin_code=list_prod.asin_code,
                image_url=list_prod.image_url,
                price=list_prod.price,
                rating=list_prod.rating,
                review_count=list_prod.review_count,
            )

    def _auto_retry_missing(
        self, enriched: list[EnrichedProduct]
    ) -> None:
        """
        Re-scrape detail pages for products with missing critical fields,
        then patch the enriched list in place.
        """
        # Collect ASINs with at least one missing field
        retry_asins: list[tuple[int, str, str]] = []
        for idx, prod in enumerate(enriched):
            for field in self._RETRY_FIELDS:
                val = getattr(prod, field, None)
                if val is None or str(val).strip() == "":
                    retry_asins.append((idx, prod.asin_code, field))
                    break

        if not retry_asins:
            self._logger.info("  Auto-retry: no missing fields found")
            return

        # Deduplicate by ASIN (one retry per product, not per field)
        unique: dict[str, int] = {}
        for idx, asin, _ in retry_asins:
            if asin not in unique:
                unique[asin] = idx

        self._logger.info(
            f"  Auto-retry: {len(unique)} products have missing fields, "
            f"re-scraping detail pages..."
        )

        for asin, idx in unique.items():
            try:
                url = enriched[idx].url
                if not url:
                    url = f"https://www.amazon.com/dp/{asin}"
                detail = self._detail_scraper.scrape(url, asin)

                # Patch missing fields
                prod = enriched[idx]
                if (not prod.bought_in_past_month or str(prod.bought_in_past_month).strip() == "") and detail.bought_in_past_month:
                    prod.bought_in_past_month = detail.bought_in_past_month
                if (not prod.rating or str(prod.rating).strip() == "") and detail.rating:
                    prod.rating = detail.rating
                if (not prod.review_count or str(prod.review_count).strip() == "") and detail.review_count:
                    prod.review_count = detail.review_count
                if (not prod.sub_category_rank or str(prod.sub_category_rank).strip() == "") and detail.sub_category_rank:
                    prod.sub_category_rank = detail.sub_category_rank
                if (not prod.brand or str(prod.brand).strip() == "") and detail.brand:
                    prod.brand = detail.brand
                if (not prod.product_size or str(prod.product_size).strip() == "") and detail.product_size:
                    prod.product_size = detail.product_size
                if (not prod.product_weight or str(prod.product_weight).strip() == "") and detail.product_weight:
                    prod.product_weight = detail.product_weight
                if not prod.has_coupon and detail.has_coupon:
                    prod.has_coupon = detail.has_coupon
                if (not prod.coupon_text or str(prod.coupon_text).strip() == "") and detail.coupon_text:
                    prod.coupon_text = detail.coupon_text

                self._logger.info(f"    [{asin}] patched: "
                                  f"bought={detail.bought_in_past_month}, "
                                  f"rating={detail.rating}, "
                                  f"brand={detail.brand}")
            except Exception as e:
                self._logger.warning(f"    [{asin}] auto-retry failed: {e}")

    def scrape_category(
        self,
        url: str,
        max_list_products: int = 100,
        max_detail_products: Optional[int] = None,
        max_reviews: int = 10,
    ) -> str | None:
        """
        Scrape one category: list page → detail pages → CSV.

        Args:
            url: The category / New Releases page URL.
            max_list_products: Max products to scrape from list page (default 100).
            max_detail_products: Max products to visit detail pages (default = max_list_products).
            max_reviews: Max reviews per product.

        Returns:
            Output CSV path, or None if no products found.
        """
        if max_detail_products is None:
            max_detail_products = max_list_products

        self._logger.info(f"\n{'='*60}")
        self._logger.info(f"Processing category: {url}")

        # Step 1: Scrape list page
        self._logger.info("Step 1: Scraping list page...")
        list_products, sub_cat_name = self._list_scraper.scrape_list_page(url, max_list_products)

        if not list_products:
            self._logger.warning("No products found on list page.")
            return None

        self._logger.info(
            f"Step 1 complete: {len(list_products)} products, "
            f"sub_category='{sub_cat_name}'"
        )

        # Step 2: Enrich with detail pages
        self._logger.info(
            f"Step 2: Scraping detail pages for up to {max_detail_products} products..."
        )
        enriched: list[EnrichedProduct] = []

        detail_count = min(len(list_products), max_detail_products)
        for idx, list_prod in enumerate(list_products[:detail_count], 1):
            self._logger.info(f"  [{idx}/{detail_count}] ASIN: {list_prod.asin_code}")
            enriched_prod = self._enrich_product(list_prod, max_reviews)
            enriched.append(enriched_prod)

            # Progress log every 10 products
            if idx % 10 == 0:
                self._logger.info(
                    f"  Progress: {idx}/{detail_count} detail pages done."
                )

        # Step 2b: Auto-retry missing fields
        if self._auto_retry:
            self._auto_retry_missing(enriched)

        # Step 3: Save outputs
        if self._output_format == "xlsx":
            xlsx_path = self._build_output_path(sub_cat_name, "xlsx")
            self._save_to_xlsx(enriched, xlsx_path)
            self._logger.info(f"Step 3: Saved XLSX to {xlsx_path}")
            return xlsx_path
        else:
            csv_path = self._build_output_path(sub_cat_name, "csv")
            self._save_to_csv(enriched, csv_path)
            self._logger.info(f"Step 3: Saved CSV to {csv_path}")
            return csv_path

    def _save_to_csv(self, products: list[EnrichedProduct], output_path: str) -> None:
        rows = [p.model_dump() for p in products]
        df = pd.DataFrame(rows)

        # Ensure image_url_display uses HYPERLINK formula for Excel image preview
        if "image_url" in df.columns and "image_url_display" not in df.columns:
            df["image_url_display"] = df["image_url"].apply(
                lambda url: f'=HYPERLINK("{url}","[图片]")' if url else ""
            )

        # Remove duplicate columns (e.g., review_count.1 from model_dump/pandas read-back)
        dup_cols = [c for c in df.columns if c.endswith(".1")]
        if dup_cols:
            df = df.drop(columns=dup_cols)

        # Build dynamic column order (includes bullet_point_N and review_N fields)
        base_cols = [
            "list_rank", "title", "url", "asin_code", "image_url_display",
            "price", "rating", "review_count",
            "brand", "bought_in_past_month", "has_coupon", "coupon_text",
            "product_size", "product_weight",
            "sub_category_name", "sub_category_rank", "has_a_plus",
        ]
        bp_cols = sorted(c for c in df.columns if c.startswith("bullet_point_"))
        review_cols = sorted(c for c in df.columns if c.startswith("review_"))
        col_order = base_cols + bp_cols + review_cols

        existing = [c for c in col_order if c in df.columns]
        df = df[existing]
        df.to_csv(output_path, index=False, encoding="utf-8-sig")

    @staticmethod
    def _rating_to_stars(val: str) -> str:
        """Convert '4.0 out of 5 stars' → '⭐⭐⭐⭐'."""
        if not val:
            return ""
        try:
            rating = float(str(val).split(" ")[0])
            stars = "".join("⭐" for _ in range(int(round(rating))))
            return stars
        except Exception:
            return val

    def _save_to_xlsx(self, products: list[EnrichedProduct], output_path: str) -> None:
        """Save to XLSX with embedded product images in column B."""
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            self._logger.warning("openpyxl not installed, skipping XLSX export")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Amazon Products"

        rows = [p.model_dump() for p in products]
        df = pd.DataFrame(rows)

        # Remove duplicate columns
        dup_cols = [c for c in df.columns if c.endswith(".1")]
        if dup_cols:
            df = df.drop(columns=dup_cols)

        # Define column order (B = image, skip image_url col)
        display_cols = [
            "list_rank", "title", "asin_code",
            "price", "rating", "review_count",
            "brand", "bought_in_past_month", "has_coupon", "coupon_text",
            "product_size", "product_weight",
            "sub_category_name", "sub_category_rank", "has_a_plus",
        ]
        bp_cols = sorted(c for c in df.columns if c.startswith("bullet_point_"))
        # Interleave: review_1_rating, review_1_text, review_2_rating, review_2_text ...
        review_cols = []
        for i in range(1, 20):
            r_col = f"review_{i}_rating"
            t_col = f"review_{i}_text"
            if r_col in df.columns:
                review_cols.append(r_col)
            if t_col in df.columns:
                review_cols.append(t_col)
            if r_col not in df.columns and t_col not in df.columns:
                break
        display_cols = display_cols + bp_cols + review_cols
        display_cols = [c for c in display_cols if c in df.columns]

        # Write header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["排名", "标题", "ASIN", "价格", "评分", "评论数",
                   "品牌", "月销量", "优惠券", "优惠券详情", "产品尺寸", "产品重量",
                   "细分类目", "分类排名", "A+页面"] \
                + [f"描述{i}" for i in range(1, 6)] \
                + [item for i in range(1, 6) for item in [f"评论{i}评分", f"评论{i}内容"]]
        # Write headers: A1 = 排名图片, B onwards = display_cols headers
        cell_a1 = ws.cell(row=1, column=1, value="排名图片")
        cell_a1.fill = header_fill
        cell_a1.font = header_font
        cell_a1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_a1.border = border

        for col_idx, header in enumerate(headers[:len(display_cols)], 2):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Set column widths
        ws.column_dimensions["A"].width = 8   # image thumbnail
        ws.column_dimensions["B"].width = 5   # rank number
        ws.column_dimensions["C"].width = 45  # title
        ws.column_dimensions["D"].width = 15  # ASIN
        for col_idx in range(5, len(display_cols) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        self._logger.info(f"  Downloading {len(products)} product images...")
        img_row_height = 60  # pixels for image row height

        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            asin = str(row.get("asin_code", ""))
            image_url = row.get("image_url", "")

            # Download image in background (lazy)
            img_path = self._download_image(asin, image_url) if image_url else None

            # Set row height
            ws.row_dimensions[row_idx].height = img_row_height

            # Column A: embedded image
            if img_path and os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    img.width = 80
                    img.height = 80
                    img.anchor = f"A{row_idx}"
                    ws.add_image(img)
                    img_cell = ws.cell(row=row_idx, column=1, value="")
                    img_cell.border = border
                except Exception as e:
                    self._logger.warning(f"    Failed to embed image for {asin}: {e}")
                    ws.cell(row=row_idx, column=1, value="[图片]").border = border
            else:
                ws.cell(row=row_idx, column=1, value=image_url[:30] if image_url else "").border = border

            # Remaining columns
            for col_offset, col_name in enumerate(display_cols, 2):
                val = row.get(col_name, "")
                if pd.isna(val):
                    val = ""
                # Convert review rating text to stars
                if col_name.endswith("_rating"):
                    val = self._rating_to_stars(str(val))
                    cell_font = Font(size=24)
                else:
                    cell_font = Font(size=9)
                cell = ws.cell(row=row_idx, column=col_offset, value=str(val) if val is not None else "")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                cell.font = cell_font

        # Freeze top row
        ws.freeze_panes = "B2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(display_cols) + 2)}{len(df) + 1}"

        wb.save(output_path)
        self._logger.info(f"  XLSX saved: {output_path} ({len(products)} products)")
